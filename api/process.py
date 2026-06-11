import json, base64, os, io
from datetime import datetime

ROW_FILTERS = [
    {'pol': 'VNHPH',             'pod': 'KR',                       'polList': ['VNHPH'],                                                         'podList': None},
    {'pol': 'CNSHK',             'pod': 'KR',                       'polList': ['CNSHK'],                                                         'podList': None},
    {'pol': 'CNXMN',             'pod': 'KR',                       'polList': ['CNXMN'],                                                         'podList': None},
    {'pol': 'VNSGN',             'pod': 'KR',                       'polList': ['VNSGN'],                                                         'podList': None},
    {'pol': 'TH',                'pod': 'KR',                       'polList': ['TH'],                                                            'podList': None},
    {'pol': 'JP',                'pod': 'KR',                       'polList': ['JPYOK','JPTYO','JPNGO','JPSMZ','JPOSA','JPUKB','JPHKT','JPMOJ'], 'podList': None},
    {'pol': 'CNSHA',             'pod': 'KRINC/KRKUV',              'polList': ['CNSHA'],                                                         'podList': ['KRINC','KRKUV']},
    {'pol': 'CNNGB',             'pod': 'KRINC/KRKUV',              'polList': ['CNNGB'],                                                         'podList': ['KRINC','KRKUV']},
    {'pol': 'CNSHA',             'pod': 'KRPUS/KRKAN/KRUSN',        'polList': ['CNSHA'],                                                         'podList': ['KRPUS','KRKAN','KRUSN']},
    {'pol': 'CNNGB',             'pod': 'KRPUS/KRKAN/KRUSN',        'polList': ['CNNGB'],                                                         'podList': ['KRPUS','KRKAN','KRUSN']},
    {'pol': 'CNTAO/CNLYG',       'pod': 'KR',                       'polList': ['CNTAO','CNLYG'],                                                 'podList': None},
    {'pol': 'CNTXG',             'pod': 'KRINC/KRPUS/KRKAN/KRUSN',  'polList': ['CNTXG'],                                                        'podList': ['KRINC','KRPUS','KRKAN','KRUSN']},
    {'pol': 'CNTXG',             'pod': 'KRPTK',                    'polList': ['CNTXG'],                                                         'podList': ['KRPTK']},
    {'pol': 'CNDLC',             'pod': 'KR',                       'polList': ['CNDLC'],                                                         'podList': None},
    {'pol': 'CNYNT',             'pod': 'KR',                       'polList': ['CNYNT'],                                                         'podList': None},
    {'pol': 'CNNKG',             'pod': 'KR',                       'polList': ['CNNKG'],                                                         'podList': None},
    {'pol': 'CNZJG',             'pod': 'KRPUS/KRKAN/KRUSN',        'polList': ['CNZJG'],                                                         'podList': ['KRPUS','KRKAN','KRUSN']},
    {'pol': 'CNNTG/CNTAG/CNZJG', 'pod': 'KRINC/KRPTK',             'polList': ['CNNTG','CNTAG','CNZJG'],                                         'podList': ['KRINC','KRPTK']},
    {'pol': 'RU',                'pod': 'KR',                       'polList': ['RUVVO','RUVFP'],                                                  'podList': None},
]

ROW_EXCEL = {
    'VNHPH|KR':                        (5,6),
    'CNSHK|KR':                        (9,10),
    'CNXMN|KR':                        (13,14),
    'VNSGN|KR':                        (17,18),
    'TH|KR':                           (21,22),
    'JP|KR':                           (25,26),
    'CNSHA|KRINC/KRKUV':               (29,30),
    'CNNGB|KRINC/KRKUV':               (33,34),
    'CNSHA|KRPUS/KRKAN/KRUSN':         (37,38),
    'CNNGB|KRPUS/KRKAN/KRUSN':         (41,42),
    'CNTAO/CNLYG|KR':                  (45,46),
    'CNTXG|KRINC/KRPUS/KRKAN/KRUSN':  (49,50),
    'CNTXG|KRPTK':                     (53,54),
    'CNDLC|KR':                        (57,58),
    'CNYNT|KR':                        (61,62),
    'CNNKG|KR':                        (65,66),
    'CNZJG|KRPUS/KRKAN/KRUSN':        (69,70),
    'CNNTG/CNTAG/CNZJG|KRINC/KRPTK': (73,74),
    'RU|KR':                           (77,78),
}

def parse_raw(wb_raw):
    ws = wb_raw.active
    headers = [str(c.value).strip() if c.value else '' for c in ws[4]]
    def idx(name, default):
        try: return headers.index(name)
        except: return default
    i_pol  = idx('POL',    13)
    i_pod  = idx('POD',    14)
    i_teu  = idx('TEU',    20)
    i_pc   = idx('P/C',    21)
    i_sttl = idx('S.TTL1', 29)
    data = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        try:
            pol  = str(row[i_pol]  or '').strip()
            pod  = str(row[i_pod]  or '').strip()
            teu  = float(row[i_teu]  or 0)
            pc   = str(row[i_pc]   or '').strip().upper()
            sttl = float(row[i_sttl] or 0)
            if pol and pol not in ('nan','POL') and teu > 0:
                data.append({'pol':pol,'pod':pod,'teu':teu,'sttl':sttl,'pc':'PP' if pc=='P' else 'CLT'})
        except:
            continue
    return data

def get_teu_rpb(data, pol_list, pod_list, pc):
    sub = [r for r in data if r['pol'] in pol_list]
    if pod_list: sub = [r for r in sub if r['pod'] in pod_list]
    if pc == 'PP':  sub = [r for r in sub if r['pc'] == 'PP']
    if pc == 'CLT': sub = [r for r in sub if r['pc'] == 'CLT']
    teu = sum(r['teu'] for r in sub)
    inc = sum(r['sttl'] for r in sub)
    return round(teu), (inc/teu if teu > 0 else 0)

def handler(request):
    if request.method != 'POST':
        return Response(json.dumps({'error':'Method not allowed'}), status=405,
                        headers={'Content-Type':'application/json'})
    try:
        import openpyxl
        from supabase import create_client

        body      = json.loads(request.body)
        file_data = body.get('fileData')
        label     = body.get('label', 'result')
        if not file_data:
            return Response(json.dumps({'error':'fileData 없음'}), status=400,
                            headers={'Content-Type':'application/json'})

        supabase = create_client(os.environ['SUPABASE_URL'], os.environ['SUPABASE_SERVICE_KEY'])

        # RAW DATA 파싱
        raw_bytes = base64.b64decode(file_data)
        wb_raw    = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        data      = parse_raw(wb_raw)
        total_teu = sum(r['teu'] for r in data)

        # 템플릿 로드 (가장 최근 파일) — keep_links=True 로 서식 보존
        file_list = supabase.storage.from_('Templates').list('', {'sortBy':{'column':'created_at','order':'desc'}})
        if not file_list:
            raise Exception('Templates 버킷에 파일이 없습니다.')
        latest    = file_list[0]['name']
        tpl_bytes = supabase.storage.from_('Templates').download(latest)
        wb = openpyxl.load_workbook(io.BytesIO(tpl_bytes), keep_links=True)
        ws = wb['RPB']

        # INBOUND TOTAL PP/CLT 행 AK/AL 함수 (없을 경우만)
        sum_pp  = '+'.join([f'AK{r}' for r in [5,9,13,17,21,25,29,33,37,41,45,49,53,57,61,65,69,73,77,81]])
        sum_clt = '+'.join([f'AK{r}' for r in [6,10,14,18,22,26,30,34,38,42,46,50,54,58,62,66,70,74,78,82]])
        if ws.cell(85,37).value is None:
            ws.cell(85,37).value = f'={sum_pp}'
            ws.cell(85,38).value = '=IFERROR(AW85/AK85,"0")'
        if ws.cell(86,37).value is None:
            ws.cell(86,37).value = f'={sum_clt}'
            ws.cell(86,38).value = '=IFERROR(AW86/AK86,"0")'

        # 데이터 입력 (AK=col37, AL=col38)
        for f in ROW_FILTERS:
            pp_row, clt_row = ROW_EXCEL[f"{f['pol']}|{f['pod']}"]
            for row_num, pc in [(pp_row,'PP'),(clt_row,'CLT')]:
                teu, rpb = get_teu_rpb(data, f['polList'], f['podList'], pc)
                ws.cell(row_num, 37).value = teu
                ws.cell(row_num, 38).value = round(rpb, 10)

        # 결과 파일 생성
        out_buf = io.BytesIO()
        wb.save(out_buf)
        out_bytes = out_buf.getvalue()

        # Supabase 저장
        today    = datetime.now().strftime('%Y-%m-%d')
        ts       = int(datetime.now().timestamp() * 1000)
        filename = f'RPB_{label}_{today}_{ts}.xlsx'
        supabase.storage.from_('Results').upload(
            filename, out_bytes,
            {'content-type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'}
        )
        url = supabase.storage.from_('Results').get_public_url(filename)

        # 히스토리
        supabase.table('rpb_history').insert({
            'filename':filename,'label':label,'url':url,
            'total_teu':int(total_teu),'created_at':datetime.now().isoformat()
        }).execute()

        return Response(json.dumps({
            'success':True,'filename':filename,'url':url,
            'fileBase64':base64.b64encode(out_bytes).decode(),
            'totalTeu':int(total_teu),
        }), status=200, headers={'Content-Type':'application/json'})

    except Exception as e:
        import traceback
        return Response(json.dumps({'error':str(e),'trace':traceback.format_exc()}),
                        status=500, headers={'Content-Type':'application/json'})
